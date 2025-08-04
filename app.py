from flask import Flask, render_template, request
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)
EXCEL_FILE = 'leads.xlsx'

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/submit', methods=['POST'])
def submit():
    name = request.form['name']
    phone = request.form['phone']
    email = request.form.get('email', '')
    location = request.form['location']

    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(['Name', 'Phone', 'Email', 'Location'])
    else:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

    ws.append([name, phone, email, location])
    wb.save(EXCEL_FILE)

    return render_template('index.html', message="Lead submitted successfully!")

if __name__ == '__main__':
    app.run(debug=True)